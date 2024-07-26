from django.shortcuts import render,redirect,HttpResponse
from .forms import SubjectForm,ReportcardForm,Excelform
from django.views import View
from .models import Student,Reportcard,Subject
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.contrib import messages
import openpyxl
from openpyxl import Workbook






# Create your views here.

# --------------------------------------- Subjects  -------------------------------

class SubjectListView(View):
    def get(self, request):
        subjects = Subject.objects.all()
        return render(request, 'subject/sub_list.html', {'subjects': subjects})
    
    
class SubjectCreateView(View):
    def get(self, request):
        form = SubjectForm()
        return render(request, 'subject/subject_form.html', {'form': form})
    
    def post(self, request):
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('subject_list')
        return render(request, 'subject/subject_form.html', {'form': form})
    

class SubjectUpdateView(View):
    def get(self, request, pk):
        subject = Subject.objects.get(pk=pk)
        form = SubjectForm(instance=subject)
        return render(request, 'subject/subject_update.html', {'form': form})
    
    def post(self, request, pk):
        subject = Subject.objects.get(pk=pk)
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            return redirect('subject_list')
        

class SubjectDeleteView(View):
    def get(self, request, pk):
        subject = Subject.objects.get(pk=pk)
        subject.delete()
        return redirect('subject_list')


# --------------------------------------- Report Card  -------------------------------   

class ReportCardListView(View):
    def get(self, request):
        reportcards = Reportcard.objects.all()
        return render(request, 'reportcard/report_list.html', {'reportcards': reportcards})
    
    
class ReportCardCreateView(View):
    def get(self, request):
        form = ReportcardForm()
        return render(request, 'reportcard/report_form.html', {'form': form})
    
    def post(self, request):
        form = ReportcardForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('reportcard_list')
        return render(request, 'reportcard/report_form.html', {'form': form})
    
class ReportCardUpdateView(View):
    def get(self, request, pk):
        reportcard = Reportcard.objects.get(pk=pk)
        form = ReportcardForm(instance=reportcard)
        return render(request, 'reportcard/report_form.html', {'form': form, 'reportcard': reportcard})
    
    def post(self, request, pk):
        reportcard = Reportcard.objects.get(pk=pk)
        form = ReportcardForm(request.POST, instance=reportcard)
        if form.is_valid():
            form.save()
            return redirect('reportcard_list')
        return render(request, 'reportcard/report_form.html', {'form': form, 'reportcard': reportcard})
    

class ReportCardDeleteView(View):
    def get(self, request, pk):
        reportcard = Reportcard.objects.get(pk=pk)
        reportcard.delete()
        return redirect('reportcard_list')



class ReportCardView(View):
    def get(self, request, pk):
        reportcard = Reportcard.objects.get(pk=pk)
        context = {
            'student_name': reportcard.stuname.first_name,
            'total_marks': reportcard.total_marks(),
            'total_grade':reportcard.total_grade(),
            'performance': reportcard.performance(),
        }
        return render(request, 'reportcard/reportcard.html', context)



# --------------------------------------- Course Completion Certificate  -------------------------------
  

class StudentListView(View):
    def get(self, request):
        students = Student.objects.all()
        context = {'students': students}
        return render(request, 'student/studentlist.html', context)
    


class GenerateCertificateView(View):
    def get(self, request, student_id):
        student = Student.objects.get(id=student_id)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{student.first_name}_{student.last_name}_certificate.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        p.setFont("Helvetica", 12)
        
        # Draw the certificate content

        # p.drawString(200, height - 50, "Course Completion Certificate")
        # p.drawString(100, height - 100, f"This is to certify that {student.first_name} {student.last_name} ")
        # p.drawString(100, height - 150, f" has successfully completed the {student.course.course_name} course.")
        
        #the above fourline code is enough for certificate generation. below code is to just for more styling

        p.setFont("Helvetica-Bold", 24)
        p.drawCentredString(width / 2.0, height - 100, "Certificate of Completion")
        p.setFont("Helvetica", 18)
        p.drawCentredString(width / 2.0, height - 150, "This is to certify that")
        p.setFont("Helvetica-Bold", 20)
        p.drawCentredString(width / 2.0, height - 200, f"{student.first_name} {student.last_name}")
        p.setFont("Helvetica", 18)
        p.drawCentredString(width / 2.0, height - 250, "has successfully completed the course")
        p.setFont("Helvetica-Bold", 20)
        p.drawCentredString(width / 2.0, height - 300, f"{student.course.course_name}")
        p.setFont("Helvetica", 16)
        p.drawCentredString(width / 2.0, height - 450, "Congratulations!")
        # Add a border (optional)
        p.rect(30, 30, width - 60, height - 60)

        # Save the PDF file
        p.showPage()
        p.save()
        
        return response
    

# --------------------------------------- Import and Export  -------------------------------

class ImportDataView(View):
    def get(self, request):
        form = Excelform()
        return render(request, 'excel/excel_upload.html', {'form': form})

    def post(self, request):
        form = Excelform(request.POST, request.FILES)
        if form.is_valid():
            if 'file' not in request.FILES:
                messages.error(request, "No file uploaded!")
                return render(request, 'excel/excel_upload.html', {'form': form})

            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                Reportcard.objects.create(
                    stuname=Student.objects.get(first_name=row[0]),  # Assuming stuname is a ForeignKey to a Student model
                    sub1=Subject.objects.get(name=row[1]),           # Assuming sub1 is a ForeignKey to a Subject model
                    mark1=row[2],
                    sub2=Subject.objects.get(name=row[3]),
                    mark2=row[4],
                    sub3=Subject.objects.get(name=row[5]),
                    mark3=row[6],
                    sub4=Subject.objects.get(name=row[7]),
                    mark4=row[8],
                    sub5=Subject.objects.get(name=row[9]),
                    mark5=row[10]
                )
            messages.success(request, "Data imported successfully!")
            return HttpResponse("imported successfully")
        return render(request, 'excel/excel_upload.html', {'form': form})


class ExportToExcelView(View):

    def get(self, request, *args, **kwargs):
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Excel Upload Data"
        # Define the headers
        headers = ['stuname', 'sub1', 'mark1','sub2','mark2','sub3','mark3','sub4','mark4','sub5','mark5']
        ws.append(headers)
        # Fetch the data from the database
        excel_uploads = Reportcard.objects.all()
        # Append the data to the worksheet
        for upload in excel_uploads:
            ws.append([upload.stuname.first_name, upload.sub1.name,upload.mark1,upload.sub2.name,upload.mark2,upload.sub3.name,upload.mark3,upload.sub4.name,upload.mark4,upload.sub5.name,upload.mark5])
        # Create an HTTP response with the appropriate headers for Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=excelupload_data.xlsx'
        # Save the workbook to the response
        wb.save(response)
        return response



    